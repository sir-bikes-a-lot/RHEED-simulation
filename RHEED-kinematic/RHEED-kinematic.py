# -*- coding: utf-8 -*-
"""
Created on Mon Apr 28 12:49:33 2025

@author: sschaefe
"""

import copy
import csv
import itertools as it
import math
import matplotlib as mpl
import matplotlib.colors as colors
import matplotlib.pyplot as plt
from numba import jit, prange
import numpy as np
import openpyxl as xl
import os
from pymatgen.core import Lattice, Structure, Molecule
from pymatgen.io.vasp import Poscar
import scipy as sp


def basistrans(R, gamma, psi):
    # basistrans
    # Transform from lattice basis vectors (a,b,c) to Cartesian basis vectors
    # (x,y,z).
    # Currently only set up for transformation in the (a,b) plane.
    # Inputs:
    # R         Vector of (a,b,c) coordinates
    # psi       Azimuth angle in (a,b) plane, in radians
    # gamma     Angle between basis vectors a and b
    # Outputs:
    # Rt        Vector of Cartesian coordinates (x,y,z)
    A = [[np.cos(psi), np.cos(gamma - psi), 0], [-np.sin(psi), np.sin(gamma - psi), 0], [0, 0, 1]]
    Rt = np.matmul(A, R)
    return Rt


def calcF(crystal, mu, S, r, k0z):
    # calcF()
    # Calculate the kinematically diffracted amplitude based on scattering
    # factors, reciprocal space vectors, and atomic positions.
    # Assume z = 0 corresponds to a "mean plane" surface, with atomic positions
    # z < 0 below the surface.
    # Inputs:
    # f         Atomic scattering factors. Vector of length m
    # mu        Effective absorption coefficient, in Angstroms^-1
    # S         Reciprocal space vectors. 3D array
    # r         Atomic positions. Array with length m and height 3
    # k0z       z-component (transverse) of incident wavevector, in Ang^-1
    #
    # Outputs:
    # A         Electron diffraction amplitude. 3D array

    # Initialize loop variables.
    m, n = S.shape[0:2]
    M = int(len(crystal['atoms']))
    tempA = np.zeros((m, n, M), dtype=complex)

    # Loop on each reciprocal space vector.
    for i in range(m):
        temp = np.zeros((n, M), dtype=complex)
        for j in range(n):
            # Loop on each atom.
            for k in range(M):
                # Calculate the diffracted amplitude.
                temp[j, k] = crystal['f'][crystal['atoms'][k]][i, j]*np.exp(1j*(S[i, j, 0]*r[k, 0] + S[i, j, 1]*r[k, 1] + S[i, j, 2]*r[k, 2]))*np.exp(mu*r[k, 2]/2)
        tempA[i] = temp

    for k in range(M):
        # If Sz is less than -k0z, the beam is totally internally reflected.
        tempA[S[:, :, 2] <= -k0z, k] = 0

    A = np.sum(tempA, 2)
    return A


@jit(parallel=True)
def calcFpar(labels, fGa, fN, mu, S, r, k0z):
    # calcFpar()
    # Calculate the kinematically diffracted amplitude based on scattering
    # factors, reciprocal space vectors, and atomic positions.
    # Assume z = 0 corresponds to a "mean plane" surface, with atomic positions
    # z < 0 below the surface.
    # Use Numba jit to parallelize loop on each atom.
    # Inputs:
    # labels    NumPy array of strings, 'Ga' or 'N', for each atom
    # fGa       NumPy array of Ga scattering factors as a function of S
    # fN        NumPy array of N scattering factors as a function of S
    # mu        Effective absorption coefficient, in Angstroms^-1
    # S         NumPy array of reciprocal space vectors. 3D array
    # r         NumPy array of atomic positions, in Angstroms
    # k0z       z-component (transverse) of incident wavevector, in Ang^-1
    #
    # Outputs:
    # A         Electron diffraction amplitude. 3D array

    # Initialize loop variables.
    m, n = S.shape[0:2]
    M = labels.shape[0]
    tempA = np.zeros((m, n, M), dtype=np.complex64)

    # Loop on each reciprocal space vector.
    for i in prange(m):
        temp = np.zeros((n, M), dtype=np.complex64)
        for j in range(n):
            # Loop on each atom.
            for k in range(M):
                # Switch on the atom type.
                if labels[m] == 'Ga':
                    fc = fGa[i, j]
                elif labels[m] == 'N':
                    fc = fN[i, j]
                else:
                    pass
                    # raise ValueError('Invalid atom label')

                # Calculate the diffracted amplitude.
                temp[j, k] = fc*np.exp(1j*(S[i, j, 0]*r[k, 0] + S[i, j, 1]*r[k, 1] + S[i, j, 2]*r[k, 2]))*np.exp(mu*r[k, 2]/2)
        tempA[i] = temp

    # If Sz is less than -k0z, the beam is totally internally reflected.
    for i in prange(m):
        for j in prange(n):
            if S[i, j, 2] <= -k0z:
                tempA[i, j, :] = 0

    A = np.sum(tempA, 2)
    return A


def calck0(E):
    # calck0
    # Calculate electron beam wavevector k0.
    # Inputs:
    # E         Electron beam energy, in eV
    #
    # Ouputs:
    # k0        Electron beam wavevector, in Ang^-1

    hbar = 1.0546e-34       # J*s
    m0 = 9.1094e-31         # kg
    c = 3e8                 # m/s
    eV = 1.6022e-19         # J = kg*m^2/s^2
    k0 = 1/hbar*np.sqrt(2*m0*E*eV + ((E*eV)**2)/(c**2))*1e-10    # Ang^-1
    return k0


def calcmu(crystal, elements, K0, delE, E):
    # calcmu
    # Calculate the mean absorption coefficient of the crystal using the
    # Compton incoherent scattering functions.
    #
    # Inputs:
    # crystal   Structure containing the crystal supercell model parameters
    # elements  Library of elemental parameters
    # K0        Incident electron beam wavevector, in Angstroms^-1
    # delE      Energy loss between incident and scattered electrons, in eV
    # E         Incident electron beam energy, in eV
    #
    # Outputs:
    # mu0       Mean absorption coefficient, in Angstroms^-1
    # Get the inelastic scattering cross-section for each atom in the crystal.
    # Do this by identifying unique instances of atom names and calculating the
    # inelastic scattering cross-section only for those atoms.
    a = []
    SigInelTot = 0
    SigInel = {}

    for m in range(len(crystal['atoms'])):
        # Check the list to see if we have calculated the scattering factors
        # yet.
        if not crystal['atoms'][m] in a:
            # Calculate inelastic scattering cross-section at each scattering
            # vector.
            SigInelc = calcSigInel(crystal['atoms'][m], elements, K0, delE, E)    # Angstroms^2

            # Pack this value into a structure.
            # SigInel = {crystal['atoms'][m]: SigInelc}
            SigInel[crystal['atoms'][m]] = SigInelc

            # Add the atom name to a list.
            if not a:
                a = crystal['atoms'][m]
            else:
                a = [a, crystal['atoms'][m]]

        # Sum up the inelastic scattering cross-sections for all the atoms in
        # the crystal.
        SigInelTot = SigInelTot + SigInel[crystal['atoms'][m]]  # Angstroms^2

    # Calculate the mean absorption coefficient by dividing by the supercell
    # volume.
    mu0 = SigInelTot/crystal['omega']    # Angstroms^-1
    return mu0


def calcSigInel(atom, elements, K0, delE, E):
    # calcSigInel
    # Calculate the inelastic scattering cross-section of the atom using the
    # Thomas-Fermi screening length formulation.
    #
    # Inputs:
    # atom      String containing the atom name
    # elements  Library of elemental parameters
    # K0        Incident electron beam wavevector, in Angstroms^-1
    # delE      Energy loss between incident and scattered electrons, in eV
    # E         Incident electron beam energy, in eV
    #
    # Outputs:
    # SigInel   Inelastic scattering cross-section of the atom, in Angstroms^2
    # Fetch the incoherent scattering function from the database of elements.
    element = elements[atom]

    # Define some physical constants.
    aH = 5.29177e-11            # m

    # Calculate integration limits Smin and Kn + K0.
    Smin = (K0/2)*(delE/E)      # Angstroms^-1

    # Use the Thomas-Fermi screening length formulation to calculate the
    # inelastic scattering cross sections.
    a = 0.885*aH*element['Z']**(-1/3)*1e10                      # Angstroms
    Sa = 1/a                                                    # Angstroms^-1
    lam = 2*np.pi/K0                                            # Angstroms
    Sig = (lam**2*element['Z']**2*a**2)/(2*np.pi*aH**2*1e20)    # Angstroms^2
    SigInel = Sig*(4/element['Z'])*np.log(Sa/Smin)              # Angstroms^2
    return SigInel


def calcSmesh(r, d, thetai, phii, k0, Nx, Ny):
    # calcSmesh
    # Setup a mesh in reciprocal space sitting on the Ewald sphere with radius
    # k0. Do this by working backwards from the RHEED screen size r.
    #
    # Inputs:
    # r         RHEED screen radius, in cm
    # d         Distance from sample to RHEED screen
    # thetai    Incident RHEED beam angle, in radians
    # phii      Incident RHEED beam azimuth with respect to x, in radians
    # k0        Incident electron beam wavevector, in Angstroms^-1
    # Nx        Number of mesh points in x direction
    # Ny        Number of mesh points in y direction
    #
    # Outputs:
    # xd, yd    2D arrays of RHEED screen coordinates, in cm
    # S         2D mesh of reciprocal space coordinates (Sx, Sy, Sz), in
    #           Angstroms^-1
    # Begin by setting up the RHEED screen coordinates. Make a symmetric mesh
    # with a point at (0,0). Use an odd number of points.
    if not Nx % 2:
        Nx += 1

    if not Ny % 2:
        Ny += 1

    # Square mesh
    x = np.linspace(-r, r, Nx)
    y = np.linspace(0, 2*r, Ny)
    xd, yd = np.meshgrid(x, y)

    # Calculate the origin of the Ewald sphere, adopting the convention that
    # the tip of the incident wavevector ends at (0,0,0).
    Sx0 = -k0*np.cos(thetai)*np.cos(phii)
    Sy0 = -k0*np.cos(thetai)*np.sin(phii)
    Sz0 = k0*np.sin(thetai)

    # Loop on each mesh point.
    S = np.zeros((Nx, Ny, 3))

    for n in range(Ny):
        # Calculate Sz.
        S[:, n, 2] = k0*(np.sin(thetai) + yd[:, n]/np.sqrt(yd[:, n]**2 + d**2))
        # Angstroms^-1

        for m in range(Nx):
            # Solve quadratic equation for Sx.
            a = 1 + (xd[m, n]/d)**2
            b = 2*(k0*np.cos(phii)*(xd[m, n]/d)**2 - Sx0 - (xd[m, n]/d)*(k0*np.sin(phii) + Sy0))
            c = Sx0**2 + (xd[m, n]/d)**2*k0**2*(np.cos(phii))**2 - 2*(xd[m, n]/d)*k0*np.cos(phii)*(k0*np.sin(phii) + Sy0) + (k0*np.sin(phii) + Sy0)**2 + (S[m, n, 2] - Sz0)**2 - k0**2

            # Take positive solution branch only.
            S[m, n, 0] = (-b + np.sqrt(b**2 - 4*a*c))/(2*a)     # Angstroms^-1

            # Solve for Sy.
            S[m, n, 1] = (xd[m, n]/d)*(S[m, n, 0] + k0*np.cos(phii)) - k0*np.sin(phii)     # Angstroms^-1
    return xd, yd, S


def DebyeWaller(u11Params, u33Params, M, T):
    # DebyeWaller
    # Calculate Debye-Waller factor B from the fitting parameters for wurtzite
    # compound semiconductors. Based on M. Schowalter, A. Rosenauer,
    # J. T. Titantah, and D. Lamoen, "Temperature-dependent Debye–Waller
    # factors for semiconductors with the wurtzite-type structure", Acta Cryst.
    # (2009). A65, 227–231.
    #
    # Inputs:
    # u11Params         Fitting parameters for atomic displacements in the a=b
    #                   direction.
    # u33Params         Fitting parameters for atomic displacements in the c
    #                   direction.
    # M                 Atomic mass, in Daltons
    # T                 Temperature, in Kelvin
    #
    # Outputs:
    # B                 Debye-Waller factor, in Angstroms^2
    # Define some physical constants.
    hbar = 1.05457e-34          # J-s = kg*m^2*s^-1
    kB = 1.380649e-23           # J/K = kg*m^2*s^-1*K^-1
    Da = 1.660539e-27           # kg

    # Unpack the fitting parameters.
    sigma11 = u11Params['sigma']
    A11 = u11Params['A']
    B11 = u11Params['B']

    sigma33 = u33Params['sigma']
    A33 = u33Params['A']
    B33 = u33Params['B']

    # Calculate the static correlation functions from the fitting parameters.
    num = 1/np.tanh(hbar/(2*kB*T)*(A11*np.exp(-(T**2)/sigma11**2) + B11))     # dimensionless
    den = A11*np.exp(-(T**2)/sigma11) + B11     # Hz
    u11 = hbar/(2*M*Da)*num/den*(1e20)          # Angstroms^2

    num = 1/np.tanh(hbar/(2*kB*T)*(A33*np.exp(-(T**2)/sigma33**2) + B33))     # dimensionless
    den = A33*np.exp(-(T**2)/sigma33) + B33     # Hz
    u33 = hbar/(2*M*Da)*num/den*(1e20)          # Angstroms^2

    # Calculate the isotropic static correlation function and convert to B.
    usq = 2*u11 + u33               # Angstroms^2
    B = 8*(np.pi**2)*usq            # Angstroms^2
    return B


def DoyleTurner(DT, B, s):
    # DoyleTurner
    # Calculate the complex electron scattering factor for scattering angle
    # theta and electron wavelength lambda. Calculate the Debye-Waller
    # correction factor in another function.
    #
    # Inputs:
    # DoyleTurner   Structure containing the Doyle-Turner parameters for the
    #               complex electron scattering factor.
    # B             Debye-Waller factor, in Ang^2
    # s             Array of scattering vectors, in Angstroms^-1
    #
    # Outputs:
    # f             Complex electron scattering factor
    # Unpack the Doyle-Turner parameters.
    aRe = [DT['a1'], DT['a2'], DT['a3'], DT['a4'], DT['a5']]    # Angstroms
    bRe = [DT['b1'], DT['b2'], DT['b3'], DT['b4'], DT['b5']]    # Angstroms^2
    aIm = [DT['aTDS1'], DT['aTDS2'], DT['aTDS3'], DT['aTDS4'], DT['aTDS5']]     # Angstroms
    bIm = [DT['bTDS1'], DT['bTDS2'], DT['bTDS3'], DT['bTDS4'], DT['bTDS5']]     # Angstroms^2

    # Calculate the complex electron scattering factor.
    f = np.zeros((len(aRe), s.shape[0], s.shape[1]), dtype=complex)

    for i in range(len(aRe)):
        f[i, :, :] = aRe[i]*np.exp(-bRe[i]*s**2) + 1j*aIm[i]*np.exp(-(bIm[i] - B/2)*s**2)

    f = sum(f, 0)*np.exp(-B*s**2)
    return f


def elements(element, material, folder='C:/Users/sschaefe/OneDrive - NREL/Characterization/RHEED/Kinematic model/Libraries/'):
    # ELEMENTS
    # Get the parameters for a specified element. Return a structure of
    # the following form:
    #
    # params = struct(  'Z', Z,...                      Atomic mass
    #                   'DoyleTurner', DoyleTurner,...  A structure containing
    #                                                   the Doyle-Turner
    #                                                   coefficients.
    #                   'u11Params', u11Params,...      A structure containing
    #                                                   the static correlation
    #                                                   function fit parameters
    #                   'u33Params', u33Params);        A structure containing
    #                                                   the static correlation
    #                                                   function fit parameters
    #
    # The real part of the Doyle-Turner coefficients is taken from Table 3 in:
    # L.M. Peng, "Electron atomic scattering factors and scattering potentials
    # of crystals", Micron 30 (1999) 625–648
    #
    # The imaginary part of the Doyle-Turner coefficients is taken from Tables
    # 3-4 in:
    # S.L. Dudarev, L.-M. Peng, M.J. Whelan, "On the Doyle-Turner
    # representation of the optical potential for RHEED calculations", Surface
    # Science 330 (1995) 86-100
    #
    # The static correlation function fit parameters are taken from Tables 3
    # and 4 in:
    # M. Schowalter, A. Rosenauer, J. T. Titantah, and D. Lamoen, "Temperature-
    # dependent Debye–Waller factors for semiconductors with the wurtzite-type
    # structure", Acta Cryst. (2009). A65, 227–231.
    # Pull in the elemental scattering data from the spreadsheet.
    wb = xl.load_workbook(folder + 'elements.xlsx')
    T1 = wb['scattering']

    # Pull in the atomic form factor and incoherent scattering function data
    # from the spreadsheet.
    T2 = wb[element]

    # Correct the NaNs in the Doyle-Turner factor table.
    M = T1.max_row
    N = T1.max_column
    for i in range(1, M + 1):
        for j in range(2, N + 1):
            if not T1.cell(i, j).value:
                T1.cell(i, j).value = 0

    T3 = wb['u11']
    T4 = wb['u33']

    # Now get just the element specified in the input argument.
    # Get the row in Table 1 corresponding to the element.
    for col in T1.iter_cols(max_col=1):
        for cell in col:
            if cell.value == element:
                T1row = T1[cell.row]

    # Unpack the lists from table 2.
    S = []
    for row in T2.iter_rows(min_col=1, max_col=1, min_row=2):
        for cell in row:
            S.append(cell.value)
    FF = []
    for row in T2.iter_rows(min_col=2, max_col=2, min_row=2):
        for cell in row:
            FF.append(cell.value)
    Inc = []
    for row in T2.iter_rows(min_col=3, max_col=3, min_row=2):
        for cell in row:
            Inc.append(cell.value)

    # Find the rows in Table 3 corresponding to the binary material.
    T3rows = []
    for col in T3.iter_cols(max_col=1):
        for cell in col:
            if cell.value == material:
                T3rows.append(T3[cell.row])
    # Now find just the row corresponding to the element.
    for i in range(len(T3rows)):
        if T3rows[i][1].value == element:
            T3row = T3rows[i]

    # Find the rows in Table 4 corresponding to the binary material.
    T4rows = []
    for col in T4.iter_cols(max_col=1):
        for cell in col:
            if cell.value == material:
                T4rows.append(T4[cell.row])
    # Now find just the row corresponding to the element.
    for i in range(len(T4rows)):
        if T4rows[i][1].value == element:
            T4row = T4rows[i]

    # Build the element parameters structure.
    DoyleTurner = {
        'a1': T1row[3].value,
        'a2': T1row[4].value,
        'a3': T1row[5].value,
        'a4': T1row[6].value,
        'a5': T1row[7].value,
        'b1': T1row[8].value,
        'b2': T1row[9].value,
        'b3': T1row[10].value,
        'b4': T1row[11].value,
        'b5': T1row[12].value,
        'aTDS1': T1row[13].value,
        'aTDS2': T1row[14].value,
        'aTDS3': T1row[15].value,
        'aTDS4': T1row[16].value,
        'aTDS5': T1row[17].value,
        'bTDS1': T1row[18].value,
        'bTDS2': T1row[19].value,
        'bTDS3': T1row[20].value,
        'bTDS4': T1row[21].value,
        'bTDS5': T1row[22].value}
    Incoherent = {
        'S': S,
        'FF': FF,
        'Inc': Inc}
    u11Params = {
        'sigma': T3row[2].value,
        'A': T3row[3].value,
        'B': T3row[4].value}
    u33Params = {
        'sigma': T4row[2].value,
        'A': T4row[3].value,
        'B': T4row[4].value}
    params = {
        'Z': T1row[1].value,
        'M': T1row[2].value,
        'DoyleTurner': DoyleTurner,
        'Incoherent': Incoherent,
        'u11Params': u11Params,
        'u33Params': u33Params}
    return params


def GaussFun(x, x0, sig):
    # GaussFun
    # Set up a Gaussian distribution W normalized such that sum(W) = 1.
    # y-offset = 0.
    #
    # Inputs
    # x         1-D vector of input values
    # x0        x-intercept
    # sig       Standard deviation
    #
    # Outputs
    # W         1-D Gaussian function

    W = np.exp(-((x - x0)**2)/(2*sig**2))
    W = W/sum(W)
    return W


def GetG(lattice, hkl):
    # GetG
    # Calculate the reciprocal lattice vectors G for each h,k,l triplet.
    #
    # Inputs:
    # lattice       3x3 array with the unit cell lattice vectors, in Cartesian
    #               coordinates. In Angstroms
    # hkl           Nx3 array of N different h,k,l triplets
    #
    # Outputs:
    # G             Nx3 array of reciprocal lattice vectors in Cartesian
    #               coordinates. In Angstroms^-1
    # Gmag          Length N vector of magnitude of reciprocal lattice vectors.

    # Calculate the reciprocal lattice vectors.
    a = lattice[0]                   # Angstroms
    b = lattice[1]                   # Angstroms
    c = lattice[2]                   # Angstroms
    omega = np.dot(a, np.cross(b, c))        # Angstroms^3

    a1 = 2*np.pi/omega*np.cross(b, c)         # Angstroms^-1
    b1 = 2*np.pi/omega*np.cross(c, a)         # Angstroms^-1
    c1 = 2*np.pi/omega*np.cross(a, b)         # Angstroms^-1

    # Generate the array of reciprocal lattice vectors Ghkl.
    N = len(hkl)
    Ghkl = np.zeros((N, 3, 3))

    i = 0
    for n in hkl:
        Ghkl[i, 0, :] = a1*n[0]
        Ghkl[i, 1, :] = b1*n[1]
        Ghkl[i, 2, :] = c1*n[2]
        i += 1

    # Flatten Ghkl into an Nx3 array of Cartesian coordinates. Calculate the
    # magnitude of G.
    G = np.zeros((N, 3))
    Gmag = np.zeros((N, 1))

    for n in range(N):
        G[n] = np.sum(Ghkl[n], axis=0)
        Gmag[n] = np.sqrt(np.dot(G[n], G[n]))
    return G, Gmag


def GetScatterFact(atom, ElementsLib, s, T):
    # GetScatterFact
    # Calculate complex electronic scattering factor for the specified atom.
    #
    # Inputs:
    # Atom          String containing the atom name.
    # elements      Library of elemental parameters
    # s             Array of scattering vectors, in Angstroms^-1
    # T             Temperature, in Kelvin
    #
    # Outputs:
    # f             Complex electron scattering factor
    # s             (Optional) scattering parameter, in Angstroms^-1
    #
    # Fetch the Doyle-Turner and static correlation function coefficients from
    # the database of elements.
    element = ElementsLib[atom]
    DT = element['DoyleTurner']
    u11Params = element['u11Params']
    u33Params = element['u33Params']

    # Calculate the Debye-Waller factor B. Neglect the anisotropy and calculate
    # the isotropic factor.
    B = DebyeWaller(u11Params, u33Params, element['M'], T)

    # Calculate the complex electron scattering factor from the Doyle-Turner
    # parameters.
    f = DoyleTurner(DT, B, s)
    return f


def GetTopSurf(r, Nx, Ny):
    # GetTopSurf
    # Calculate a surface corresponding to the uppermost layer of atoms in the
    # crystal model. Do this by binning the atoms in the x-y plane and picking
    # off the highest atoms in the z-direction.
    #
    # Inputs:
    # r             NumPy array of Cartesian x-y-z coordinates
    # Nx            Number of x-bins to use
    # Ny            Number of y-bins to use
    #
    # Outputs:
    # z0            z-coordinate of upper surface for each atom
    # Xedges        (Optional) edges of x-bins
    # Yedges        (Optional) edges of y-bins
    # zmax          (Optional) z-coordinates of top surface

    # Get the crystal coordinates.
    rx = r[:, 0]
    ry = r[:, 1]
    rz = r[:, 2]
    M = len(rx)

    # Set up edges of bins to span the entire x-y plane of the sample.
    Xedges = np.linspace(min(rx), max(rx), Nx)
    Yedges = np.linspace(min(ry), max(ry), Ny)

    # Initialize the array of maximum heights of the surface. If a height is
    # not binned, then assign it the minimum value.
    zmax = min(rz)*(np.zeros((Nx, Ny)) + 1)

    # Perform binning along the x-direction.
    ryBin = []
    rzBin = []
    xbins = np.digitize(rx, Xedges)
    xind = np.unique(xbins)
    for i in xind:
        ryBin.append([ry[xbins == i]])
        rzBin.append([rz[xbins == i]])

    # Loop on each of the x-binned coordinates and perform y-binning.
    for i in range(len(rzBin)):
        rzBin2 = []
        # Get the array of binned values for the i-th x-bin.
        rzB = np.array(rzBin[i])

        # Perform binning along the y-direction.
        ybins = np.digitize(ryBin[i], Yedges)
        yind = np.unique(ybins)
        for j in yind:
            rzBin2.append([rzB[ybins == j]])

        # Loop on each of the y-binned coordinates and pick off the max value.
        for j in range(len(rzBin2)):
            zmax[xind[i]-1, yind[j]-1] = np.max(np.array(rzBin2[j]))

    # Add this max height to the crystal structure for each atom.
    # Re-bin the atoms along the y-direction.
    ybins = np.digitize(ry, Yedges)
    z0 = np.zeros((M, 1))

    # Loop through each atom and assign it the maximum z height based on its
    # x-y bin.
    for m in range(M):
        z0[m] = zmax[xbins[m]-1, ybins[m]-1]
    return z0, Xedges, Yedges, zmax


def innerPot(K0, theta, psi, S, Vi):
    # innerPot
    # Calculate the internal scattering vectors after applying boundary
    # conditions to the incident electron beam.
    #
    # Inputs:
    # K0        Electron beam wavevector in vacuum, in Ang^-1
    # theta     Incident beam angle, in radians
    # psi       Incident beam azimuth, in radians
    # S         Array of vacuum scattering vector triplets, in Ang^-1
    # Vi        Inner potential in crystal, in eV
    #
    # Outputs:
    # s         Array of internal scattering vector triplets, in Ang^-1
    # k0z       z-component (transverse) of incident wavevector, in Ang^-1

    # Define some physical constants.
    hbar = 1.0546e-34   # J*s
    m0 = 9.1094e-31     # kg
    eV = 1.6022e-19     # J = kg*m^2/s^2

    # Unpack the vacuum scattering vector.
    Sx = S[:, :, 0]
    Sy = S[:, :, 1]
    Sz = S[:, :, 2]

    # Calculate the x, y, and z-components of the vacuum wavevector.
    # K0x = K0*np.cos(theta)*np.cos(psi)   # Angstroms^-1
    # K0y = K0*np.cos(theta)*np.sin(psi)   # Angstroms^-1
    K0z = -K0*np.sin(theta)         # Angstroms^-1

    # Calculate the z-component of the scattered electron wavevector in vacuum.
    Kz = Sz + K0z                   # Angstroms^-1

    # Calculate the z-component of the scattered wavevector in the crystal.
    kz = np.sqrt((Kz*1e10)**2 + 2*m0*eV*Vi/(hbar**2))*(1e-10)    # Angstroms^-1
    k0z = np.sqrt((K0z*1e10)**2 + 2*m0*eV*Vi/(hbar**2))*(1e-10)  # Angstroms^-1

    # Calculate the z-component of the internal scattering vector.
    sz = kz - k0z                  # Angstroms^-1

    # Pack up the internal scattering vector. The transverse x- and
    # y-components are unchanged from vacuum.
    s = np.zeros(np.shape(S))
    s[:, :, 0] = Sx
    s[:, :, 1] = Sy
    s[:, :, 2] = sz

    # Pack up the internal scattered wavevector. The transverse x- and
    # y-components are unchanged from vacuum.
    # k = zeros(size(S));
    # k[:, :, 0] = Sx + K0x
    # k[:, :, 1] = Sy + K0y
    # k[:, :, 2] = kz
    return s, k0z


def Kikuchi(lattice, K0, Vi, n):
    # Kikuchi
    # Solve for the location of the Kikuchi lines.
    #
    # Inputs:
    # lattice       3x3 array with the unit cell lattice vectors, in Cartesian
    #               coordinates. In Angstroms
    # K0            Incident electron beam wavevector, in Angstroms^-1
    # Vi            Inner potential, in eV
    # n             Max index for h,k,l triplets.
    #
    # Outputs:
    # kk            N x Nx x 3 array of scattered wavevector triplets
    #               corresponding to the Kikuchi lines.
    # Set up the array of h,k,l triplets up to n.
    v = list(range(-n, n+1))                # Integers between -n and n
    p = []
    for i in range(3):
        p.extend(v)                         # 3 of each integer in this vector
    hkl = list()
    for comb in it.combinations(p, 3):      # Permute the integer triplets
        hkl.extend([comb])
    hkl = set(hkl)                          # Keep the unique tuples
    # Debug
    # hkl = {(1, 1, 1)}
    N = len(hkl)

    # Get the array of reciprocal lattice vectors.
    G, Gmag = GetG(lattice, hkl)

    # Define some physical constants.
    hbar = 1.0546e-34       # J*s
    m0 = 9.1094e-31         # kg
    eV = 1.6022e-19         # J = kg*m^2/s^2

    # The wavevector magnitude in the crystal is increased by the inner
    # potential Vi.
    k = np.sqrt(K0**2 + 2*m0*eV*Vi/(hbar**2)*(1e-20))       # Angstroms^-1

    # Set up a vector of internal scattered wavevectors kx to solve for the
    # Kikuchi lines.
    kx = np.linspace(0, K0, 10001)    # Angstroms^-1

    # Solve for the Kikuchi lines, which satisfy k.G = (|G|^2)/2.
    kk1 = np.zeros((N, len(kx), 3))
    kk2 = np.zeros((N, len(kx), 3))

    for m in range(N):
        # Define auxiliary variables a, b, and c to solve the quadratic
        # equation.
        a = G[m, 1]**2 + G[m, 2]**2
        b = -2*G[m, 1]*((Gmag[m, 0]**2)/2 - kx*G[m, 0])
        c = ((Gmag[m, 0]**2)/2 - kx*G[m, 0])**2 - (G[m, 2]**2)*(k**2 - kx**2)
        ky1 = (-b + np.sqrt(b**2 - 4*a*c))/(2*a)
        ky2 = (-b - np.sqrt(b**2 - 4*a*c))/(2*a)

        # If the value of ky is imaginary, assign it the value 'NaN'.
        ky1t = ky1[~np.isnan(ky1)]
        ky2t = ky2[~np.isnan(ky2)]

        # Solve for kz.
        kz1 = np.sqrt(k**2 - kx[~np.isnan(ky1)]**2 - ky1t**2)
        kz2 = np.sqrt(k**2 - kx[~np.isnan(ky2)]**2 - ky2t**2)
        # kz1 = kz1[~np.isnan(kz1)]
        # kz2 = kz2[~np.isnan(kz2)]

        # Pack up the scattered wavevector.
        kk1[m, 0:len(ky1t), 0] = kx[~np.isnan(ky1)]
        kk1[m, 0:len(ky1t), 1] = ky1t
        kk1[m, 0:len(ky1t), 2] = kz1

        kk2[m, 0:len(ky2t), 0] = kx[~np.isnan(ky2)]
        kk2[m, 0:len(ky2t), 1] = ky2t
        kk2[m, 0:len(ky2t), 2] = kz2
    return kk1, kk2


def LoadCrystal(filename, hkl, ElementsLib, s, T, h=1e10):
    # LoadCrystal
    # Load a VASP POSCAR file containing the crystal information, including the
    # atom IDs and coordinates. Uses the VASPLAB package: https://github.com/max-radin/VASPLAB
    #
    # Inputs:
    # filename      String containing the file name. Automatically detect the
    #               file type and switch.
    # hkl           Triplet with supercell size, integers >= 1.
    # ElementsLib   Library of elemental parameters
    # s             Array of scattering vectors, in Angstroms^-1
    # T             Temperature, in Kelvin
    # h             Cut-off height/depth for crystal slab in z-direction.
    #               Measured in Angstroms from the maximum z-position in the
    #               crystal model.
    #
    # Outputs:
    # crystal       Structure containing the atom names and Cartesian
    #               coordinates (x,y,z), the atomic scattering factors f,
    #               and the lattice.
    # Parse the filename.
    name, ext = os.path.splitext(filename)

    # Switch based on file extension. VASP POSCAR file
    if ext == '.vasp':
        # Load the .vasp file.
        # geometry = Poscar.from_file(filename)
        # structure = geometry.structure
        geometry = Structure.from_file(filename)
        geometryExt = copy.deepcopy(geometry)

        # Generate a supercell of size (h,k,l).
        geometryExt = geometryExt.make_supercell(hkl)

        # Unpack the lattice basis vectors and atom names.
        lattice = geometryExt.lattice.matrix
        atomnames = list(geometryExt.symbol_set)

        # Get a list of the atom names.
        atoms0 = np.array(geometryExt.labels)

        # Get the Cartesian coordinates of each atom in the supercell
        r0 = geometryExt.cart_coords

        # Offset all the z-coordinates such that z=0 corresponds to the
        # highest fractional coordinate.
        z0 = max(r0[:, 2])
        r0[:, 2] = r0[:, 2] - z0

        # Get the upper surface of the atomic coordinates.
        Nx = 101
        Ny = 101
        z0, Xedges, Yedges, zmax = GetTopSurf(r0, Nx, Ny)

        # Get the crystal z-coordinate for each atom.
        rz = r0[:, 2]

        # Scrape the crystal.
        atoms = atoms0[rz >= z0[:, 0]-h]
        r = r0[rz >= z0[:, 0]-h, :]

        # Calculate the supercell volume.
        omega = sp.integrate.trapezoid(sp.integrate.trapezoid(zmax, x=Yedges), x=Xedges)
        omega = omega - sp.integrate.trapezoid(sp.integrate.trapezoid(zmax - h, x=Yedges), x=Xedges)

        # Calculate atomic scattering factor at each reciprocal space mesh
        # point for each element in the lattice.
        f = {}
        for m in range(len(atomnames)):
            # Calculate scattering factor at each scattering vector.
            # DEBUG: Turn off s-dependence.
            # s = 0*s
            fc = GetScatterFact(atomnames[m], ElementsLib, s, T)

            # Pack this array into a structure.
            f[atomnames[m]] = fc

            # Pack up the atom names and coordinates to the output structure.
        crystal = {
            'atoms': atoms,
            'r': r,
            'z0': z0,
            'f': f,
            'lattice': lattice,
            'UClattice': geometry.lattice.matrix,
            'omega': omega}

    # Dump file from MD simulation.
    else:
        # Load the dump file.
        atoms0, r0 = ParseDump(filename)

        # Offset all the z-coordinates such that z=0 corresponds to the
        # highest fractional coordinate.
        z0 = max(r0[:, 2])
        r0[:, 2] = r0[:, 2] - z0

        # Assign the supercell lattice parameters.
        lattice = [[max(r0[:, 0]), 0, 0],
                   [0, max(r0[:, 1]), 0],
                   [0, 0, -min(r0[:, 2])]]

        # Calculate atomic scattering factor at each reciprocal space mesh
        # point for each element in the lattice.
        atomnames = np.unique(atoms0)
        f = {}
        for m in range(len(atomnames)):
            # Calculate scattering factor at each scattering vector.
            # DEBUG: Turn off s-dependence.
            # s = 0*s
            fc = GetScatterFact(atomnames[m], ElementsLib, s, T)

            # Pack this array into a structure.
            f[atomnames[m]] = fc

        # Get the upper surface of the atomic coordinates. Use Nx x Ny mesh
        # points.
        Nx = 101
        Ny = 101
        z0, Xedges, Yedges, zmax = GetTopSurf(r0, Nx, Ny)

        # Get the crystal z-coordinate for each atom.
        rz = r0[:, 2]

        # Scrape the crystal to get only the atoms lying between the top
        # surface and a depth h below this surface.
        atoms = atoms0[rz >= z0[:, 0]-h]
        r = r0[rz >= z0[:, 0]-h, :]

        # Calculate the supercell volume.
        omega = sp.integrate.trapezoid(sp.integrate.trapezoid(zmax, x=Yedges), x=Xedges)
        omega = omega - sp.integrate.trapezoid(sp.integrate.trapezoid(zmax - h, x=Yedges), x=Xedges)

        # Pack up the atom names and coordinates to the output structure.
        crystal = {
            'atoms': atoms,
            'r': r,
            'z0': z0,
            'f': f,
            'lattice': lattice,
            'UClattice': lattice,
            'omega': omega}
    return crystal, Xedges, Yedges, zmax


def ParseDump(filename):
    # ParseDump()
    # Read in a dump file from molecular dynamics simulation and extract the
    # atomic labels and Cartesian (x,y,z) coordinates.

    # Open the file.
    with open(filename, 'r', newline='') as f:
        # Get the first 8 columns of the file.
        subset = [row[0:7] for row in csv.reader(f, delimiter=' ')]
        # The atom IDs are listed in the second column.
        atom_ID = np.array([col[1] for col in subset[9:]])
        # The Cartesian coordinates are listed in the third - fifth columns.
        r = np.array([col[2:5] for col in subset[9:]], dtype=np.float_)
        # The coordination number is lsted in the seventh column.
        CN = np.array([col[6] for col in subset[9:]], dtype =np.int_)

    # Filter out all atoms that are not part of the film.
    atom_ID = atom_ID[CN.astype(int) > 1]
    r = r[CN.astype(int) > 1]

    # Convert from the atom ID to element name, Ga or N.
    atoms = ['']*len(atom_ID)
    bool1 = atom_ID.astype(int) == 1
    bool2 = atom_ID.astype(int) == 4
    bool3 = atom_ID.astype(int) == 2
    bool4 = atom_ID.astype(int) == 5
    for i in range(len(atoms)):
        if bool1[i] or bool2[i]:
            atoms[i] = "Ga"
        elif bool3[i] or bool4[i]:
            atoms[i] = "N"

    for i in range(len(atoms)):
        if not atoms[i]:
            raise ValueError('Error. Unexpected atom type encountered. No element assigned.')
    atoms = np.array(atoms)
    return atoms, r


def ProjScreen(k, K0, Vi, thetai, psii, d):
    # ProjScreen
    # Project the internal scattered wavevectors k onto the RHEED screen.
    # Account for the inner potential Vi.
    #
    # Inputs:
    # k             Nx3 array of internal scattered wavevectors in Cartesian
    #               coordinates. In Angstroms^-1
    # K0            Incident electron beam wavevector, in Angstroms^-1
    # Vi            Inner potential, in eV
    # thetaii       Incident beam angle, in radians
    # psii          Incident beam azimuth, in radians
    # d             Distance from sample to RHEED screen, in cm
    #
    # Outputs:
    # x, y          Screen coordinates, in cm
    # Define some physical constants.
    hbar = 1.0546e-34   # J*s
    m0 = 9.1094e-31     # kg
    eV = 1.6022e-19     # J = kg*m^2/s^2

    # Calculate the x, y, and z components of the incident electron beam.
    K0x = K0*np.cos(thetai)*np.cos(psii)
    K0y = K0*np.cos(thetai)*np.sin(psii)
    K0z = -K0*np.sin(thetai)

    # Calculate the external (vacuum) scattering vector S. The perpendicular
    # component is modified by the inner potential Vi.
    Sx0 = k[:, 0] - K0x
    Sy0 = k[:, 1] - K0y
    Kz0 = np.sqrt(k[:, 2]**2 - 2*m0*eV*Vi/(hbar**2)*(1e-20))
    # Keep only the values that are not NaN.
    Kz = Kz0[~np.isnan(Kz0)]
    Sz = Kz - K0z
    # Trim the Sx and Sy vectors to match the length of Sz.
    Sx = Sx0[~np.isnan(Kz0)]
    Sy = Sy0[~np.isnan(Kz0)]

    # Convert the scattering vectors to scattering angles.
    thetaf = np.arcsin(Sz/K0 - np.sin(thetai))
    psif = np.arctan((Sy/K0 + np.sin(psii))/(Sx/K0 + np.cos(psii)))

    # Project the scattering angles onto the RHEED screen.
    x = d*np.tan(psif)
    y = d*np.tan(thetaf)
    return x, y


def PlotRHEED(xd, yd, I, thresh=1e10, d=[], theta=[], KK=[], G2D=[]):
    # PlotRHEED()
    # Plot the RHEED pattern and Kikuchi lines, if specified.
    #
    # Inputs:
    # xd            2D array of x-positions on the RHEED screen, in cm
    # yd            2D array of y-positions on the RHEED screen, in cm
    # I             2D array of intensity values at each x-y point,
    #               dimensionless
    # thresh        Thresholding value for clipping RHEED intensity.
    # d             (Optional) Distance from sample to RHEED screen, for
    #               plotting specular reflection, in cm
    # theta         (Optional) Incident beam glancing angle, for plotting
    #               specular reflection, in radians
    # KK            (Optional) A structure containing the Kikuchi line
    #               x-y positions
    # G2D           (Optional) A structure containing the parameters for 2D
    #               Gaussian broadening of the RHEED pattern
    #
    # Outputs:
    # L0            The RHEED intensity after broadening but before intensity
    #               thresholding (clipping)
    # L             The RHEED intensity after broadening and intensity
    #               thresholding (clipping)

    # Assign a default value of zero (do not plot Kikuchi lines) if KK argument
    # is empty.
    if not KK:
        dokk = False
    else:
        # Otherwise, unpack the Kikuchi line parameters.
        dokk = True
        xk1 = KK['xk1']
        yk1 = KK['yk1']
        xk2 = KK['xk2']
        yk2 = KK['yk2']

    # Assign a default value of zero (do not apply 2D broadening) is G2D
    # argument is empty.
    if not G2D:
        broaden = False
    else:
        # Otherwise, unpack the broadening parameters.
        broaden = True
        xlim = G2D['xlim']
        ylim = G2D['ylim']
        sigx = G2D['sigx']
        sigy = G2D['sigy']
        Nl = G2D['Nl']

    if broaden:
        # Create the 2D Gaussian convolution function.
        H = np.zeros((Nl, Nl))
        x = np.linspace(-xlim, xlim, Nl)    # Ang^-1
        y = np.linspace(-ylim, ylim, Nl)    # Ang^-1

        for i in range(Nl):
            for j in range(Nl):
                H[i, j] = np.exp(-((x[i])**2)/(2*sigx**2) -((y[j])**2)/(2*sigy**2))

        # Broaden the reciprocal lattice points using convolution (faster).
        L0 = sp.convolve2d(I, H, 'same')
    else:
        L0 = I

    # Get the size of the x-y position arrays.
    Nx, Ny = np.shape(xd)

    # Apply a max threshold to the intensity map.
    L = L0

    for m in range(Nx):
        for n in range(Ny):
            if L0[m, n] > thresh:
                L[m, n] = thresh

    # Plot RHEED pattern as a colormap.
    fig, ax = plt.subplots()
    cmapr = mpl.colormaps.get_cmap('binary_r')
    plt.pcolormesh(xd, yd, np.real(L), cmap=cmapr)

    # Plot the specular reflection, if specified.
    if d and theta:
        plt.scatter(0, d*np.tan(theta), s=40, color=(0, 0.7, 0.7), edgecolors=(0, 0.5, 0.5), linewidths=1.5)

    # Plot the Kikuchi lines, if specified.
    if dokk:
        for m in range(len(xk1)):
            plt.plot(xk1[m, ~(xk1[m, :] == 0)], yk1[m, ~(yk1[m, :] == 0)], color='white', linewidth=0.5)
            plt.plot(xk2[m, ~(xk2[m, :] == 0)], yk2[m, ~(yk2[m, :] == 0)], color='white', linewidth=0.5)

    # Label the plot and scale axes.
    ax.set_xlabel('x$_{d}$ (cm)')
    ax.set_ylabel('y$_{d}$ (cm)')
    set_size(5, 5)      # in inches
    plt.axis([np.min(xd), np.max(xd), np.min(yd), np.max(yd)])
    plt.show()

    return L0, L


def set_size(w, h, ax=None):
    """ w, h: width, height in inches """
    if not ax:
        ax = plt.gca()
    l = ax.figure.subplotpars.left
    r = ax.figure.subplotpars.right
    t = ax.figure.subplotpars.top
    b = ax.figure.subplotpars.bottom
    figw = float(w)/(r-l)
    figh = float(h)/(t-b)
    ax.figure.set_size_inches(figw, figh)
    

def CalcRHEED(filename, theta, psi=0, T=300, radius=3.0, d=30, Nxy=(101, 101), hkl=(1, 1, 1), E0=20e3, dokk=False):
    # CalcRHEED()
    # Simulate reflection high-energy electron diffraction (RHEED) patterns for
    # arbitrary atomic arrangements using the kinematic diffraction
    # approximation.
    # Based on: Ayahiko Ichimaya and Philip I Cohen, "Reflection High Energy
    # Electron Diffraction", Chapters 4, 10, and 11, Cambridge University
    # Press (2004).
    # Inputs:
    # filename      Complete file name of VASP file, e.g. "my_crystal.vasp"
    # theta         Incident electron beam glancing angle, in radians
    # psi           Incident electron beam azimuth w.r.t x-axis direction,
    #               in radians
    # T             Temperature, in K
    # radius        Radius of RHEED screen port, in cm
    # d             Distance from sample to RHEED screen, in cm
    # Nxy           Doublet of integers [Nx, Ny] specifying the number of mesh
    #               points for the RHEED screen.
    # hkl           Triplet of integers defining supercell dimensions.
    #               E.g. (1 1 1) for a single unit cell
    # E0            Incident electron beam energy, in eV
    # kiklines      Logical value specifying whether or not to calculate
    #               Kikuchi lines (default = False, do not calculate)
    # Outputs:
    # x             2D array of x-positions on the RHEED screen, in cm
    # y             2D array of y-positions on the RHEED screen, in cm
    # I             2D array of intensity values at each x-y point,
    #               dimensionless
    # crystal       (Optional) return the structure containing all the crystal
    #               parameters
    # KK            (Optional) return a structure containing the Kikuchi line
    #               x-y positions

    # Add the file paths to the crystal models, elemental parameter libraries,
    # and VASP file functions.

    # Import parameters for each atom from library.
    Ga = elements('Ga', 'GaN')
    N = elements('N', 'GaN')
    ElementsLib = {'Ga': Ga, 'N': N}

    # Calculate electron beam wavevector.
    K0 = calck0(E0)         # Ang^-1

    # Set up RHEED screen coordinate mesh and corresponding reciprocal space
    # mesh.
    x, y, S = calcSmesh(radius, d, theta, 0, K0, Nxy[0], Nxy[1])

    # TO-DO: Implement a way to calculate the inner potential, Vi.
    # Assume a fixed value of 10 eV here.
    Vi = 10                 # eV

    # Convert the vacuum scattering vectors S to the internal scattering
    # vectors s.
    s, k0z = innerPot(K0, theta, psi, S, Vi)    # Angstroms^-1

    # Calculate magnitude of the internal scattering vectors.
    smag = np.sqrt(s[:, :, 0]**2 + s[:, :, 1]**2 + s[:, :, 2]**2)/(4*np.pi)     # Angstroms^-1

    # Estimate the absorption depth of the crystal using mu0 = 0.0042
    # Angstroms^-1.
    h = np.sin(theta)/0.0042            # Angstroms

    # Import crystal model.
    crystal, Xedges, Yedges, zmax = LoadCrystal(filename, hkl, ElementsLib, smag, T, h)

    # Loop through all the atoms in the list.
    M = len(crystal['atoms'])
    r = np.zeros((M, 3))

    for m in range(M):
        # Rotate crystal by azimuthal angle psi.
        r[m, :] = basistrans(crystal['r'][m, :], np.pi/2, psi)  # Angstroms

    # Calculate the mean absorption coefficient. Assume a fixed inelastic
    # scattering energy loss, delE.
    delE = 15                                           # eV
    mu0 = calcmu(crystal, ElementsLib, K0, delE, E0)    # Ang^-1
    mu = mu0/np.sin(theta)                              # Ang^-1

    # Perform Kikuchi line calculation if specified.
    if dokk:
        # Calculate the Kikuchi line locations.
        kk1, kk2 = Kikuchi(crystal['UClattice'], K0, Vi, 2)

        # Project the Kikuchi line wavevectors onto the RHEED screen.
        sz = np.shape(kk1)
        xk1 = np.zeros(sz[0:2])
        yk1 = np.zeros(sz[0:2])
        xk2 = np.zeros(sz[0:2])
        yk2 = np.zeros(sz[0:2])

        for m in range(sz[0]):
            x1, y1 = ProjScreen(kk1[m, :, :], K0, Vi, theta, psi, d)
            xk1[m, 0:len(x1)] = x1
            yk1[m, 0:len(y1)] = y1
            x2, y2 = ProjScreen(kk2[m, :, :], K0, Vi, theta, psi, d)
            xk2[m, 0:len(x2)] = x2
            yk2[m, 0:len(y2)] = y2

        # Pack the Kikuchi line locations up into an output structure.
        KK = {'xk1': xk1, 'yk1': yk1, 'xk2': xk2, 'yk2': yk2}
    else:
        # Do not calculate Kikuchi lines.
        KK = []

    # Calculate diffracted intensity.
    # A = calcF(crystal, mu, s, r, k0z)   # Dimensionless
    A = calcFpar(np.array(crystal['atoms']), crystal['f']['Ga'], crystal['f']['N'], mu, s, r, k0z)                  # Dimensionless
    I = A*np.conjugate(A)              # Dimensionless

    return x, y, I, crystal, KK


# filename = "C:/Users/sschaefe/OneDrive - NREL/Characterization/RHEED/Kinematic model/Crystal models/bulk_GaN.vasp"
filename = "C:/Users/sschaefe/OneDrive - NREL/Characterization/RHEED/Kinematic model/Crystal models/GaN example dumps/2873_7A_2.vasp"
theta = 3
psi = 0
hkl = (1, 1, 1)
thresh = 1e6

xd, yd, I, crystal, KK = CalcRHEED(filename, theta*np.pi/180, psi*np.pi/180, T=300, radius=3.0, d=30, Nxy=(201, 201), hkl=hkl, E0=20e3, dokk=False)

L0, L = PlotRHEED(xd, yd, I, thresh, 30, theta*np.pi/180, KK)
